#!/usr/bin/env python3
import csv
import os
import sys
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from reportlab.lib.utils import ImageReader
from PIL import Image

LOGO_PATH = os.path.join(os.path.dirname(os.path.realpath(__file__)), "logo", "images.png")
DEFAULT_CABANG = "Kotabaru"
DEFAULT_MANAGER = "Endang Komara"

_logo_cache = None

def get_logo_reader():
    global _logo_cache
    if _logo_cache is not None:
        return _logo_cache
    if os.path.exists(LOGO_PATH):
        try:
            img = Image.open(LOGO_PATH)
            if img.mode == 'P':
                img = img.convert('RGBA')
            _logo_cache = ImageReader(img)
            return _logo_cache
        except Exception:
            pass
    return None

# Try to import pandas for Excel support
try:
    import pandas as pd  # type: ignore
    # Fix for 'extLst' error in openpyxl
    from openpyxl.styles.fills import PatternFill  # type: ignore
    original_init = PatternFill.__init__
    def patched_init(self, *args, **kwargs):
        kwargs.pop('extLst', None)
        return original_init(self, *args, **kwargs)
    PatternFill.__init__ = patched_init
except ImportError:
    pd = None

def draw_spk(c, data, x_offset, width_half, height, spk_type="PENYEGELAN"):
    margin = 1.2 * cm
    line_y = height - 3.5 * cm
    
    # --- Header ---
    c.setFont("Helvetica", 10)
    c.drawCentredString(x_offset + width_half/2 + 1*cm, height - 1.5*cm, "PERUSAHAAN UMUM DAERAH AIR MINUM")
    c.setFont("Helvetica-Bold", 11)
    c.drawCentredString(x_offset + width_half/2 + 1*cm, height - 2.0*cm, "TIRTA TARUM KABUPATEN KARAWANG")
    c.setFont("Helvetica", 9)
    c.drawCentredString(x_offset + width_half/2 + 1*cm, height - 2.5*cm, "Jl. Surotokunto No.205 Karawang Timur")
    
    # Logo
    logo = get_logo_reader()
    if logo:
        c.drawImage(logo, x_offset + margin, height - 3.2*cm, width=2*cm, height=2*cm, preserveAspectRatio=True)
    
    # Thick Line
    c.setLineWidth(1.5)
    c.line(x_offset + margin, line_y, x_offset + width_half - margin, line_y)

    # --- Title & Nomor ---
    c.setFont("Helvetica-Bold", 10)
    title_text = f"SURAT PERINTAH KERJA {spk_type.upper()}"
    c.drawCentredString(x_offset + width_half/2, line_y - 1*cm, title_text)
    c.setFont("Helvetica", 10)
    c.drawCentredString(x_offset + width_half/2, line_y - 1.5*cm, "Nomor : …../…../SPK/2026")

    # --- Sender/Recipient ---
    c.setFont("Helvetica", 10)
    curr_y = line_y - 2.5*cm
    c.drawString(x_offset + margin, curr_y, f"Dari           : Manager {data.get('cabang', '')}")
    curr_y -= 0.5*cm
    c.drawString(x_offset + margin, curr_y, "Untuk        : Distribusi")
    
    curr_y -= 1*cm
    c.drawString(x_offset + margin, curr_y, f"Untuk melaksanakan pekerjaan {spk_type.upper()} sambungan pelanggan")
    curr_y -= 0.45*cm
    c.drawString(x_offset + margin, curr_y, "sebagaimana data dibawah ini")

    # --- Data Pelanggan ---
    curr_y -= 1*cm
    c.drawString(x_offset + margin + 1*cm, curr_y, f"No pelanggan  : {data.get('no_pelanggan', '')}")
    curr_y -= 0.5*cm
    c.drawString(x_offset + margin + 1*cm, curr_y, f"Nama               : {data.get('nama', '')}")
    
    curr_y -= 1*cm
    c.drawString(x_offset + margin + 1*cm, curr_y, "Rincian tunggakan air / Non air")
    curr_y -= 0.8*cm
    c.drawString(x_offset + margin + 1*cm, curr_y, f"Total Tunggakan      : {data.get('total_tunggakan_bulan', '')} Bulan")
    curr_y -= 0.5*cm
    c.drawString(x_offset + margin + 1*cm, curr_y, f"Total tagihan             : Rp.{data.get('total_tagihan', '')}")

    curr_y -= 1.2*cm
    c.drawString(x_offset + margin, curr_y, "Demikian untuk dilaksanakan dengan penuh tanggung jawab dan dengan semestinya")

    # --- Signatures ---
    curr_y -= 1.5*cm
    c.drawCentredString(x_offset + margin + 3*cm, curr_y, "Tanda tangan Pelanggan")
    c.drawCentredString(x_offset + width_half - margin - 3*cm, curr_y, f"Manager Cabang {data.get('cabang', '')}")
    
    curr_y -= 2*cm
    c.drawCentredString(x_offset + margin + 3*cm, curr_y, "(…………………………..)")
    c.setFont("Helvetica-Bold", 10)
    c.drawCentredString(x_offset + width_half - margin - 3*cm, curr_y, data.get('manager_name', ''))

    # --- Technical Data (Bottom) ---
    c.setFont("Helvetica", 9)
    tech_y = curr_y - 1.5*cm
    c.drawString(x_offset + margin, tech_y, "Merk Meter")
    c.drawString(x_offset + margin + 3.5*cm, tech_y, ":")
    c.drawRightString(x_offset + width_half - margin, tech_y, "Tanda tangan Petugas")
    
    tech_y -= 0.45*cm
    c.drawString(x_offset + margin, tech_y, "No Meter")
    c.drawString(x_offset + margin + 3.5*cm, tech_y, ":")
    
    tech_y -= 0.45*cm
    stand_label = "Stand Meter Segel" if "SEGEL" in spk_type.upper() else "Stand Meter Cabut"
    c.drawString(x_offset + margin, tech_y, stand_label)
    c.drawString(x_offset + margin + 3.5*cm, tech_y, ":")
    
    tech_y -= 0.45*cm
    segel_cabut_label = "No Segel" if "SEGEL" in spk_type.upper() else "No Body"
    c.drawString(x_offset + margin, tech_y, segel_cabut_label)
    c.drawString(x_offset + margin + 3.5*cm, tech_y, ":")
    
    tech_y -= 0.45*cm
    c.drawString(x_offset + margin, tech_y, "Digit Meter")
    c.drawString(x_offset + margin + 3.5*cm, tech_y, ":")
    c.drawRightString(x_offset + width_half - margin, tech_y - 0.5*cm, "(...................................)")

def add_customer_page(c, data, spk_type="PENYEGELAN"):
    width, height = landscape(A4)
    width_half = width / 2
    draw_spk(c, data, 0, width_half, height, spk_type)
    c.setDash(3, 3)
    c.setLineWidth(0.5)
    c.line(width_half, 0.5*cm, width_half, height - 0.5*cm)
    c.setDash(1, 0)
    draw_spk(c, data, width_half, width_half, height, spk_type)
    c.showPage()

def generate_grouped_pdf(data_list, output_path, spk_type):
    if not data_list:
        return False
    c = canvas.Canvas(output_path, pagesize=landscape(A4))
    for row in data_list:
        if str(row.get('no_pelanggan', '')).strip():
            add_customer_page(c, row, spk_type)
    c.save()
    return True

def normalize_row(row, spk_type):
    normalized = {}
    cols = {k.upper().strip(): v for k, v in row.items()}
    
    def format_no_pelanggan(val):
        if pd and pd.isna(val):
            return ''
        if val is None or str(val).strip() == '':
            return ''
        s = str(val).strip()
        if s.isdigit() and not s.startswith('0'):
            return '0' + s
        return s
    
    if 'PENYEGELAN' in spk_type.upper():
        normalized['no_pelanggan'] = format_no_pelanggan(cols.get('NOMOR PELANGGAN', cols.get('NO PELANGGAN', '')))
        normalized['nama'] = cols.get('NAMA', '')
        normalized['total_tunggakan_bulan'] = cols.get('JUMLAH BLN', cols.get('JUMLAH BULAN', ''))
        normalized['total_tagihan'] = cols.get('TOTAL REK', cols.get('TOTAL TAGIHAN', ''))
    else:
        normalized['no_pelanggan'] = format_no_pelanggan(cols.get('NO SAMB', cols.get('NO PELANGGAN', '')))
        normalized['nama'] = cols.get('NAMA', '')
        normalized['total_tunggakan_bulan'] = cols.get('TOTAL TUNGGAKAN', '')
        normalized['total_tagihan'] = cols.get('JUMLAH TUNGGAKAN (RP)', cols.get('TOTAL TAGIHAN', ''))
    
    normalized['cabang'] = cols.get('CABANG', '') or DEFAULT_CABANG
    normalized['manager_name'] = cols.get('MANAGER', '') or DEFAULT_MANAGER
    
    for k, v in normalized.items():
        if k == 'no_pelanggan':
            continue
        if pd and pd.isna(v):
            normalized[k] = ''
        elif v is None:
            normalized[k] = ''
        else:
            normalized[k] = str(v)
    
    return normalized

def load_excel_data(excel_file):
    if pd is None:
        print("Error: pandas and openpyxl are required.")
        return None
    try:
        all_data = {}
        with pd.ExcelFile(excel_file, engine='openpyxl') as xls:
            for sheet_name in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name)
                df.columns = [str(c).strip() for c in df.columns]
                df = df.fillna('')
                records = [normalize_row(row, sheet_name) for row in df.to_dict('records')]
                all_data[sheet_name] = records
        return all_data
    except Exception as e:
        print(f"Error loading Excel: {e}")
        print("\nTips: Jika error berlanjut, buka file Excel tersebut, lalu 'Save As' dengan nama baru.")
        return None

def clear_screen():
    os.system('cls' if os.name == 'nt' else 'clear')

def print_header():
    clear_screen()
    print("\033[1;36m" + "═" * 60 + "\033[0m")
    print("\033[1;36m║\033[0m" + " " * 18 + "\033[1;33m📧 MAILING LIST CLI\033[0m" + " " * 21 + "\033[1;36m║\033[0m")
    print("\033[1;36m║\033[0m" + " " * 14 + "\033[90mPDF Generator & Report Tool\033[0m" + " " * 17 + "\033[1;36m║\033[0m")
    print("\033[1;36m" + "═" * 60 + "\033[0m")

def print_menu():
    print("\n\033[1;35m┌─ MENU UTAMA\033[0m")
    print("\033[1;35m│\033[0m")
    print("\033[1;35m│\033[0m  \033[1;32m[1]\033[0m \033[90mLihat\033[0m Daftar Pelanggan (CSV)")
    print("\033[1;35m│\033[0m  \033[1;32m[2]\033[0m \033[90mCetak\033[0m Laporan dari Excel")
    print("\033[1;35m│\033[0m  \033[1;32m[3]\033[0m \033[90mCetak\033[0m Laporan dari CSV")
    print("\033[1;35m│\033[0m  \033[1;31m[q]\033[0m \033[90mKeluar\033[0m")
    print("\033[1;35m│\033[0m")

def print_status(message, status="info"):
    icons = {"info": "ℹ️", "success": "✅", "error": "❌", "warning": "⚠️"}
    colors = {"info": "34", "success": "32", "error": "31", "warning": "33"}
    icon = icons.get(status, "ℹ️")
    color = colors.get(status, "34")
    print(f"\033[1;{color}m{icon} {message}\033[0m")

def print_table(headers, rows, max_width=58):
    if not rows:
        print("\033[90m   (Tidak ada data)\033[0m")
        return
    
    col_widths = [len(h) for h in headers]
    for row in rows:
        for i, cell in enumerate(row):
            col_widths[i] = max(col_widths[i], len(str(cell)[:20]))
    
    total_width = sum(col_widths) + 3 * (len(headers) - 1) + 4
    
    print("\033[90m" + "┌" + "─" * (total_width - 2) + "┐" + "\033[0m")
    header_str = "\033[90m│\033[0m " + " \033[90m│\033[0m ".join(f"\033[1m{h}\033[0m".ljust(w) for h, w in zip(headers, col_widths)) + " \033[90m│\033[0m"
    print(header_str)
    print("\033[90m" + "├" + "─" * (total_width - 2) + "┤" + "\033[0m")
    
    for row in rows[:20]:
        row_str = "\033[90m│\033[0m " + " \033[90m│\033[0m ".join(str(c)[:20].ljust(w) for c, w in zip(row, col_widths)) + " \033[90m│\033[0m"
        print(row_str)
    
    if len(rows) > 20:
        print("\033[90m│\033[0m" + f" ... dan {len(rows) - 20} data lainnya".center(total_width - 2) + "\033[90m│\033[0m")
    
    print("\033[90m" + "└" + "─" * (total_width - 2) + "┘" + "\033[0m")

def print_progress(current, total, prefix=""):
    percent = (current / total) * 100
    filled = int(30 * current / total)
    bar = "█" * filled + "░" * (30 - filled)
    print(f"\r\033[90m{prefix}\033[0m [{bar}] \033[1;36m{percent:.1f}%\033[0m", end="", flush=True)

def interactive_mode():
    csv_file = "data_pelanggan.csv"
    output_dir = "output"
    
    # Load initial data from CSV if exists
    data = []
    if os.path.exists(csv_file):
        with open(csv_file, mode='r', encoding='utf-8') as f:
            data = list(csv.DictReader(f))

    while True:
        print_header()
        
        # Status info
        csv_status = "\033[32m✓\033[0m" if data else "\033[31m✗\033[0m"
        output_status = "\033[32m✓\033[0m" if os.path.exists(output_dir) else "\033[31m✗\033[0m"
        print(f"\033[90m  CSV Data:\033[0m {csv_status} {len(data)} records  \033[90m│\033[0m  \033[90mOutput Dir:\033[0m {output_status}")
        print()
        
        print_menu()
        
        try:
            choice = input("\033[1;36m  Pilih menu ›\033[0m ").strip().lower()
        except (KeyboardInterrupt, EOFError):
            print("\n\n\033[90m👋 Keluar...\033[0m")
            break
        
        if choice == '1':
            print_header()
            print("\n\033[1;35m┌─ DAFTAR PELANGGAN\033[0m\n")
            if data:
                rows = [[row.get('no_pelanggan', '-'), row.get('nama', '-')] for row in data]
                print_table(["No Pelanggan", "Nama"], rows)
            else:
                print_status("Data CSV kosong atau tidak ditemukan", "warning")
            input("\n\033[90m  Tekan Enter untuk kembali...\033[0m")
            
        elif choice == '2':
            print_header()
            print("\n\033[1;35m┌─ CETAK LAPORAN DARI EXCEL\033[0m\n")
            print("\033[90m  Drag & drop file Excel atau ketik path\033[0m")
            print("\033[90m  Default: PENYEGELAN & PENCABUTAN JAN 26.xlsx\033[0m\n")
            
            try:
                excel_file = input("\033[1;36m  File ›\033[0m ").strip().strip("'\"") or "PENYEGELAN & PENCABUTAN JAN 26.xlsx"
            except (KeyboardInterrupt, EOFError):
                print("\n\n\033[90m👋 Keluar...\033[0m")
                break
            
            if os.path.exists(excel_file):
                print_status(f"Membaca file: {os.path.basename(excel_file)}...", "info")
                data_dict = load_excel_data(excel_file)
                if data_dict:
                    if not os.path.exists(output_dir): 
                        os.makedirs(output_dir)
                        print_status(f"Membuat direktori: {output_dir}", "info")
                    
                    total_sheets = len(data_dict)
                    for idx, (sheet, rows) in enumerate(data_dict.items(), 1):
                        path = os.path.join(output_dir, f"SPK_{sheet}.pdf")
                        print_progress(idx, total_sheets, f"  Generating {sheet}")
                        generate_grouped_pdf(rows, path, sheet)
                    
                    print()  # New line after progress
                    print_status(f"Berhasil generate {total_sheets} file PDF!", "success")
                    print(f"\033[90m  📁 Lokasi: ./{output_dir}/\033[0m")
                else:
                    print_status("Gagal membaca data Excel", "error")
            else:
                print_status(f"File tidak ditemukan: {excel_file}", "error")
            
            input("\n\033[90m  Tekan Enter untuk kembali...\033[0m")
            
        elif choice == '3':
            print_header()
            print("\n\033[1;35m┌─ CETAK LAPORAN DARI CSV\033[0m\n")
            
            if data:
                if not os.path.exists(output_dir): 
                    os.makedirs(output_dir)
                path = os.path.join(output_dir, "SPK_PENYEGELAN.pdf")
                print_status("Generating PDF...", "info")
                generate_grouped_pdf(data, path, "PENYEGELAN")
                print_status(f"Berhasil: {path}", "success")
            else:
                print_status("Data CSV kosong atau tidak ditemukan", "warning")
            
            input("\n\033[90m  Tekan Enter untuk kembali...\033[0m")
            
        elif choice == 'q':
            clear_screen()
            print("\n\033[1;36m" + "═" * 60 + "\033[0m")
            print("\033[1;33m  ✨ Terima kasih telah menggunakan Mailing List CLI!\033[0m")
            print("\033[1;36m" + "═" * 60 + "\033[0m\n")
            break

def main():
    output_dir = "output"
    if len(sys.argv) > 1:
        input_file = sys.argv[1]
        if input_file.endswith('.csv'):
            # ... process csv ...
            pass
        elif input_file.endswith(('.xlsx', '.xls')):
            data_dict = load_excel_data(input_file)
            if data_dict:
                if not os.path.exists(output_dir): os.makedirs(output_dir)
                for sheet, rows in data_dict.items():
                    path = os.path.join(output_dir, f"SPK_{sheet}.pdf")
                    generate_grouped_pdf(rows, path, sheet)
                    print(f"Generated: {path}")
    else:
        interactive_mode()

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nKeluar...")
